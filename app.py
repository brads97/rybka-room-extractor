import streamlit as st
import anthropic
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import io
from datetime import datetime

# ========== PAGE CONFIG ==========
st.set_page_config(
    page_title="Rybka Room Data Extractor",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ========== CUSTOM CSS (RYBKA THEME) ==========
st.markdown("""
<style>
    /* Rybka Blue Theme */
    :root {
        --rybka-blue: #003D7A;
        --rybka-light-blue: #0066CC;
        --rybka-accent: #E8F1F8;
    }
    
    
    /* Section headers */
    .main h3 {
        color: #003D7A;
        font-weight: 600;
    }       
    /* Main container */
    .main {
        background-color: #F5F7FA;
    }
    
    /* Header styling */
    .rybka-header {
        background: linear-gradient(135deg, #003D7A 0%, #0066CC 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    .rybka-title {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        color: white;
    }
    
    .rybka-subtitle {
        font-size: 1.2rem;
        margin-top: 0.5rem;
        opacity: 0.9;
        color: white;
    }
    
    /* Card styling */
    .info-card {
        background: white;
        padding: 1.5rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        margin-bottom: 1rem;
        border-left: 4px solid #0066CC;
    }
    
    .info-card h3 {
        color: #003D7A;
        margin-bottom: 0.5rem;
        font-size: 1.2rem;
    }
    
    .info-card p {
        color: #4A5568;
        margin: 0;
        font-size: 0.95rem;
    }
    
    /* Button styling */
    .stButton > button {
        background: linear-gradient(135deg, #003D7A 0%, #0066CC 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        font-size: 1.1rem;
        font-weight: 600;
        border-radius: 6px;
        transition: all 0.3s ease;
        width: 100%;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(0, 61, 122, 0.3);
    }
    
    /* File uploader */
    .uploadedFile {
    border: 2px dashed #0066CC;
    border-radius: 8px;
    padding: 1rem;
    }

    /* Force file uploader text to be visible - but NOT the drag/drop area */
    [data-testid="stFileUploader"] > label > div {
        color: #1F2937 !important;
    }

    [data-testid="stFileUploader"] small {
        color: #6B7280 !important;
    }

    /* File name and size text ONLY (after upload) */
    [data-testid="stFileUploader"] [data-testid="stMarkdownContainer"] p {
        color: #1F2937 !important;
    }

    /* Uploaded file info */
    .uploadedFileName {
        color: #1F2937 !important;
    }

    .uploadedFileData {
        color: #6B7280 !important;
    }

    /* Keep drag-and-drop section text white/light */
    [data-testid="stFileUploader"] section[data-testid="stFileUploadDropzone"] {
        color: #E5E7EB !important;
    }

    [data-testid="stFileUploader"] section[data-testid="stFileUploadDropzone"] span {
        color: #E5E7EB !important;
    }

    /* General text color fix - but don't override everything */
    .main {
        color: #1F2937;
    }
    
    /* Progress styling */
    .stProgress > div > div {
        background-color: #0066CC;
    }
    
    /* Success/Error messages - FORCE visibility */
    .stSuccess {
        background-color: #D1FAE5 !important;
        border-left: 4px solid #10B981 !important;
        color: #065F46 !important;
    }

    .stSuccess * {
        color: #065F46 !important;
    }

    .stSuccess p, .stSuccess span, .stSuccess div {
        color: #065F46 !important;
        font-weight: 500 !important;
    }

    /* Target the actual success icon and text container */
    [data-testid="stNotification"] {
        color: #065F46 !important;
    }

    [data-testid="stNotification"] * {
        color: #065F46 !important;
    }
    
    .stError {
        background-color: #FADBD8;
        border-left: 4px solid #E74C3C;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem;
        color: #7F8C8D;
        font-size: 0.9rem;
        margin-top: 3rem;
    }
    /* Download button - force white text and blue background */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #003D7A 0%, #0066CC 100%) !important;
        color: #FFFFFF !important;
        border: none !important;
    }

    .stDownloadButton > button p {
        color: #FFFFFF !important;
    }

    .stDownloadButton > button span {
        color: #FFFFFF !important;
    }

    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(0, 61, 122, 0.3) !important;
    }
</style>
""", unsafe_allow_html=True)

# ========== CORE FUNCTIONS ==========

def extract_text_with_coordinates(pdf_bytes):
    """Extract all text from PDF with their coordinates."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]  # First page
    
    text_blocks = page.get_text("dict")["blocks"]
    
    extracted_items = []
    for block in text_blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        extracted_items.append({
                            "text": text,
                            "x": span["bbox"][0],
                            "y": span["bbox"][1],
                            "width": span["bbox"][2] - span["bbox"][0],
                            "height": span["bbox"][3] - span["bbox"][1]
                        })
    
    doc.close()
    return extracted_items

def extract_floor_level(text_items, client):
    """Use Claude to identify the floor level from title block text."""
    # Get all text and look for floor indicators directly first
    all_text = " ".join([item['text'] for item in text_items])
    
    # Direct pattern matching as backup
    import re
    text_lower = all_text.lower()
    
    # Try direct pattern matching first
    patterns = {
        r'ground\s*floor': 'Ground Floor',
        r'first\s*floor': 'First Floor',
        r'1st\s*floor': 'First Floor',
        r'second\s*floor': 'Second Floor',
        r'2nd\s*floor': 'Second Floor',
        r'third\s*floor': 'Third Floor',
        r'3rd\s*floor': 'Third Floor',
        r'basement': 'Basement',
        r'lower\s*ground': 'Basement',
        r'level\s*0+1\b': 'First Floor',
        r'level\s*0+2\b': 'Second Floor',
        r'level\s*0+3\b': 'Third Floor',
    }
    
    for pattern, level in patterns.items():
        if re.search(pattern, text_lower):
            return level
    
    # If pattern matching fails, use Claude
    try:
        prompt = f"""Extract the floor level from this architectural drawing text.

EXTRACTED TEXT FROM PDF:
{all_text[:3000]}

Look for phrases like:
- "Ground Floor Plan" → return "Ground Floor"
- "First Floor Plan" → return "First Floor"  
- "Second Floor" → return "Second Floor"
- "Basement Plan" → return "Basement"
- "Level 01" or "L01" → return "First Floor"
- "Level 02" or "L02" → return "Second Floor"
- etc.

CRITICAL RULES:
1. Search the entire text for floor level indicators
2. Common locations: title blocks, drawing titles, sheet names
3. Return ONLY one of these exact formats:
   - "Basement"
   - "Ground Floor"
   - "First Floor"
   - "Second Floor"
   - "Third Floor"
   - "Fourth Floor"
   - "Fifth Floor"
   (etc.)
4. If you find "Ground Floor Plan" in the text, return "Ground Floor"
5. If you cannot find ANY floor indicator, return "Unknown"

Do not explain, just return the floor level name."""

        message = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=50,
            messages=[{"role": "user", "content": prompt}]
        )
        
        floor_level = message.content[0].text.strip()
        
        # Clean up the response
        floor_level = floor_level.replace('"', '').replace("'", "").strip()
        
        return floor_level if floor_level and floor_level != "Unknown" else "Unknown"
    except Exception as e:
        st.warning(f"Could not extract floor level: {str(e)}")
        return "Unknown"

def group_text_with_claude(text_items, client):
    """Use Claude to intelligently group extracted text into room records."""
    text_list = []
    for i, item in enumerate(text_items):
        text_list.append(f"{i}: '{item['text']}' at position (x:{item['x']:.1f}, y:{item['y']:.1f})")
    
    text_summary = "\n".join(text_list)
    
    prompt = f"""You are analyzing text extracted from an architectural floor plan PDF. Below is ALL the text found in the document with their coordinates.

Your task: Group this text into room records. Each room typically has 2-3 text labels near each other (room name, space type, area).

EXTRACTED TEXT (with coordinates):
{text_summary}

STRICT RULES:
1. You can ONLY use text from the list above - you cannot add any text that isn't listed
2. Group text items that are close together spatially (similar x,y coordinates)
3. Identify which text is:
   - room_name: specific room identifier (e.g., "Classroom 05", "Store", "Pupil WC")
   - room_number: if there's a separate number/code (e.g., "05", "101", "A-23")
   - space_type: category/function (e.g., "Teaching Space", "Circulation", "Hygiene Area")  
   - area: size with m² (e.g., "56 m²", "13 m²")
4. Ignore legend text, title blocks, scale bars, and other non-room labels
5. Skip text that clearly isn't labeling a room space
6. If you cannot confidently identify what a text item represents, don't include it

QUALITY CHECKS:
- Does each room_name actually appear in the extracted text list above?
- Are you grouping text that is spatially close together?
- Did you avoid including legend categories as room names?

Return ONLY valid JSON array:
[
  {{
    "room_name": "Classroom 05",
    "room_number": "05",
    "space_type": "Teaching Space",
    "area": "56 m²"
  }}
]

If a field is unclear or not present in the text group, use null."""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        
        response_text = message.content[0].text
        
        try:
            start_idx = response_text.find('[')
            end_idx = response_text.rfind(']') + 1
            json_str = response_text[start_idx:end_idx]
            rooms_data = json.loads(json_str)
            return rooms_data
        except Exception as e:
            st.error(f"Error parsing JSON response: {e}")
            st.code(response_text[:500])  # Show first 500 chars of response
            return []
    except Exception as e:
        st.error(f"Error calling Claude API: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return []

def sort_rooms(rooms_data):
    """Sort rooms by floor level and then alphabetically by room name."""
    floor_order = {
        "Basement": 0, "Lower Ground Floor": 1, "Ground Floor": 2,
        "First Floor": 3, "Second Floor": 4, "Third Floor": 5,
        "Fourth Floor": 6, "Fifth Floor": 7, "Sixth Floor": 8,
        "Seventh Floor": 9, "Eighth Floor": 10, "Ninth Floor": 11,
        "Tenth Floor": 12, "Unknown": 999
    }
    
    def sort_key(room):
        level = room.get("level", "Unknown")
        room_name = room.get("room_name", "")
        floor_num = floor_order.get(level, 999)
        return (floor_num, room_name.lower())
    
    return sorted(rooms_data, key=sort_key)

def create_excel(rooms_data):
    """Create Excel file with ventilation template format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Data"
    
    # Title section
    ws['A1'] = 'Calculation:'
    ws['B1'] = 'Ventilation'
    ws['A2'] = 'Project Name:'
    ws['A3'] = 'Project Number:'
    ws['A4'] = 'Revision:'
    ws['A5'] = 'Date:'
    ws['A6'] = 'By:'
    ws['A7'] = 'Approved:'
    
    ws['E1'] = 'Reference Data'
    ws['E2'] = 'Ceiling Height (m)'
    ws['F2'] = 3
    ws.merge_cells('E1:F1')
    
    ws['A9'] = 'System/Zone'
    ws['L9'] = 'Total (l/s)'
    ws['O9'] = 'Total (l/s)'
    
    # Headers
    headers = [
        'Level', 'Room Name', 'Room Number', 'Room Type', 'Floor Area (m2)',
        'Strategy', 'Occupancy (No.)', 'Volume (m³)', 'Supply (l/p/s)',
        'Supply (l/p/m2)', 'Supply Calc (l/s)', 'Supply (l/s)',
        'Extract (ACH)', 'Extract Calc (l/s)', 'Extract (l/s)'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
    
    sorted_rooms = sort_rooms(rooms_data)
    start_row = 12
    
    for idx, room in enumerate(sorted_rooms):
        row_num = start_row + idx
        
        area_str = room.get("area", "")
        try:
            area_value = float(area_str.replace("m²", "").replace("m2", "").strip()) if area_str else 0
        except:
            area_value = 0
        
        ws.cell(row=row_num, column=1).value = room.get("level", "Unknown")
        ws.cell(row=row_num, column=2).value = room.get("room_name", "")
        ws.cell(row=row_num, column=3).value = room.get("room_number", "")
        ws.cell(row=row_num, column=4).value = room.get("space_type", "")
        ws.cell(row=row_num, column=5).value = area_value if area_value > 0 else ""
        ws.cell(row=row_num, column=6).value = ""
        ws.cell(row=row_num, column=7).value = ""
        
        if area_value > 0:
            ws.cell(row=row_num, column=8).value = f"=E{row_num}*$F$2"
        else:
            ws.cell(row=row_num, column=8).value = 0
        
        ws.cell(row=row_num, column=9).value = ""
        ws.cell(row=row_num, column=10).value = ""
        ws.cell(row=row_num, column=11).value = f"=MAX(I{row_num}*G{row_num},E{row_num}*J{row_num})"
        ws.cell(row=row_num, column=12).value = f"=ROUNDUP(K{row_num},0)"
        ws.cell(row=row_num, column=13).value = ""
        ws.cell(row=row_num, column=14).value = f"=M{row_num}*H{row_num}/3.6"
        ws.cell(row=row_num, column=15).value = f"=ROUNDUP(N{row_num},0)"
    
    end_row = start_row + len(sorted_rooms) - 1
    ws['L10'] = f"=SUM(L{start_row}:L{end_row})"
    ws['O10'] = f"=SUM(O{start_row}:O{end_row})"
    
    # Column widths
    widths = [12, 25, 15, 20, 16, 12, 16, 14, 14, 16, 16, 13, 15, 17, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== STREAMLIT APP ==========

def main():
    # Initialize session state
    if 'processing' not in st.session_state:
        st.session_state.processing = False
    
    # Header
    st.markdown("""
    <div class="rybka-header">
        <h1 class="rybka-title">🏢 Rybka Room Data Extractor</h1>
        <p class="rybka-subtitle">Architectural Floor Plan Analysis Tool</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Info cards
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="info-card">
            <h3>📄 Upload PDFs</h3>
            <p>Upload one or multiple architectural floor plan PDFs</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="info-card">
            <h3>🤖 AI Processing</h3>
            <p>Automatically extract room data using Claude AI</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="info-card">
            <h3>📊 Export Excel</h3>
            <p>Download formatted ventilation calculation spreadsheet</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # API Key - Hardcoded (hidden from users)
    # Replace YOUR_API_KEY_HERE with your actual Claude API key
    api_key = st.secrets.get("ANTHROPIC_API_KEY", None)
    
    # Fallback for local development - you can hardcode here temporarily
    if not api_key:
        api_key = "YOUR_API_KEY_HERE"  # Replace with your actual key for local testing
    
    # Only show config section if API key is not set
    if not api_key or api_key == "YOUR_API_KEY_HERE":
        with st.expander("⚙️ Configuration", expanded=True):
            api_key = st.text_input(
                "Claude API Key",
                type="password",
                help="Enter your Anthropic Claude API key. Get one at console.anthropic.com"
            )
            if api_key:
                st.success("✓ API Key configured")
    else:
        # API key is configured via secrets, don't show the config section
        pass
    
    # File upload
    st.markdown("### 📤 Upload Floor Plans")
    
    st.markdown("""
    <div style="background-color: #FFF3CD; padding: 1rem; border-radius: 6px; border-left: 4px solid #FFC107; margin-bottom: 1rem;">
        <p style="margin: 0; color: #856404; font-weight: 500;">⚠️ <strong>Upload Tip:</strong> For best results, upload and process files one at a time.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Reset button
    if st.button("🔄 Reset Uploader", help="Click if uploads are failing"):
        st.cache_data.clear()
        st.session_state.clear()
        st.rerun()
    
    # RESTORED: Multiple file uploader
    uploaded_files = st.file_uploader(
        "Choose PDF files",
        type=['pdf'],
        accept_multiple_files=True,
        help="Upload one or more architectural floor plan PDFs",
        key="pdf_uploader"
    )
    
    if uploaded_files and api_key:
        # Check file sizes
        oversized_files = []
        for f in uploaded_files:
            f.seek(0, 2)  # Seek to end
            size_mb = f.tell() / (1024 * 1024)
            f.seek(0)  # Reset to beginning
            if size_mb > 10:
                oversized_files.append(f"{f.name} ({size_mb:.1f}MB)")
        
        if oversized_files:
            st.error(f"❌ The following files are too large (max 10MB):\n" + "\n".join([f"- {f}" for f in oversized_files]))
            st.stop()
        
        # Show uploaded files
        if len(uploaded_files) == 1:
            st.success(f"✅ **File ready:** {uploaded_files[0].name}")
        else:
            st.success(f"✅ **{len(uploaded_files)} files ready for processing**")
            with st.expander("📋 View uploaded files"):
                for f in uploaded_files:
                    st.write(f"• {f.name}")
        
        if st.button("🚀 Extract Room Data", use_container_width=True, type="primary"):
            if not api_key or api_key == "":
                st.error("❌ API key not configured. Please contact your administrator.")
                st.stop()
                
            # Prevent double-clicking
            if st.session_state.processing:
                st.warning("⏳ Already processing... please wait")
                st.stop()
                
            st.session_state.processing = True
            
            try:
                client = anthropic.Anthropic(api_key=api_key)
                all_rooms = []
                
                # Prepare files for processing
                files_to_process = []
                for uploaded_file in uploaded_files:
                    files_to_process.append({
                        'name': uploaded_file.name,
                        'bytes': uploaded_file.read()
                    })
                
                progress_bar = st.progress(0, text="Starting extraction...")
                status_text = st.empty()
                
                for idx, file_data in enumerate(files_to_process):
                    file_progress = (idx + 1) / len(files_to_process)
                    status_text.markdown(f"**Processing:** {file_data['name']} ({idx + 1}/{len(files_to_process)})")
                    
                    try:
                        pdf_bytes = file_data['bytes']
                        
                        if len(pdf_bytes) == 0:
                            st.error(f"❌ {file_data['name']} is empty or corrupted")
                            continue
                        
                        progress_bar.progress(file_progress * 0.3, text=f"Extracting text from {file_data['name']}...")
                        text_items = extract_text_with_coordinates(pdf_bytes)
                        
                        if len(text_items) == 0:
                            st.warning(f"⚠️ No text found in {file_data['name']}")
                            continue
                        
                        # Get floor level
                        progress_bar.progress(file_progress * 0.5, text=f"Identifying floor level in {file_data['name']}...")
                        floor_level = extract_floor_level(text_items, client)
                        
                        # Group rooms
                        progress_bar.progress(file_progress * 0.8, text=f"Grouping room data in {file_data['name']}...")
                        rooms = group_text_with_claude(text_items, client)
                        
                        # Add floor level
                        for room in rooms:
                            room["level"] = floor_level
                        
                        all_rooms.extend(rooms)
                        progress_bar.progress(file_progress, text=f"Completed {file_data['name']}")
                        
                    except Exception as file_error:
                        st.error(f"❌ Error processing {file_data['name']}: {str(file_error)}")
                        import traceback
                        st.code(traceback.format_exc())
                        continue
                
                status_text.empty()
                progress_bar.empty()
                
                # Create Excel
                st.success(f"✅ Successfully extracted {len(all_rooms)} rooms from {len(files_to_process)} file(s)!")
                
                # Show preview
                st.markdown("### 📋 Preview")
                preview_data = []
                for room in all_rooms[:10]:
                    preview_data.append({
                        "Level": room.get("level", ""),
                        "Room Name": room.get("room_name", ""),
                        "Room Type": room.get("space_type", ""),
                        "Area": room.get("area", "")
                    })
                st.dataframe(preview_data, use_container_width=True)
                
                if len(all_rooms) > 10:
                    st.info(f"Showing 10 of {len(all_rooms)} rooms")
                
                # Download button
                excel_file = create_excel(all_rooms)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_file,
                    file_name=f"room_data_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
            finally:
                st.session_state.processing = False
    
    elif uploaded_files and not api_key:
        st.warning("⚠️ Please enter your Claude API key in the Configuration section above")
    
    # Footer
    st.markdown("""
    <div class="footer">
        <p>Powered by Claude AI & Rybka Building Physics | Version 1.0.0</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
